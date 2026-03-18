"""
output.py — Excel and PDF report generators for Smartify Plus Phase Zero

Excel (generate_excel):
  - Sheet "Extracted Data": full transaction table, styled header, auto column widths,
    GST colour-coded rows, red bold for low-confidence cells
  - Sheet "Summary": aggregated metrics (total amount, GST breakdown, low-confidence count)

PDF (generate_pdf_report):
  - Landscape A4 via ReportLab Platypus
  - Professional corporate header with total amount summary
  - GST breakdown summary table with colour coding
  - Full transaction detail table with:
    - Paragraph cells for proper word wrapping of long descriptions/notes
    - GST-based row background colouring (green/amber/red)
    - Red bold confidence values for low-confidence records
    - Header row repeats across pages (repeatRows=1)
  - Accountant disclaimer footer
"""

import io
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# ─── Shared constants ─────────────────────────────────────────────────────────

# Maps internal field names → human-readable column headers
_COLUMN_MAP: dict[str, str] = {
    "date":        "Date",
    "vendor":      "Vendor",
    "description": "Description",
    "amount":      "Amount ($)",
    "gst_code":    "GST Code",
    "confidence":  "Confidence",
    "notes":       "Notes",
}

# ReportLab GST row background colours (soft, print-friendly)
_GST_COLOURS: dict[str, colors.Color] = {
    "10%":     colors.Color(0.84, 0.94, 0.84),   # Soft green  — taxable supply
    "0%":      colors.Color(1.00, 0.95, 0.80),   # Soft amber  — GST-free supply
    "unknown": colors.Color(1.00, 0.88, 0.88),   # Soft red    — needs review
}

# openpyxl hex fills for GST rows
_XL_GST_FILLS: dict[str, PatternFill] = {
    "10%":     PatternFill("solid", fgColor="D4EDDA"),  # Green
    "0%":      PatternFill("solid", fgColor="FFF3CD"),  # Amber
    "unknown": PatternFill("solid", fgColor="F8D7DA"),  # Red
}

_HEADER_BLUE    = colors.HexColor("#1F4E79")   # Corporate dark blue
_XL_HEADER_HEX  = "1F4E79"                     # Same for openpyxl


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _format_currency(amount: float | None) -> str:
    """Format float as AUD currency string, or '—' for None."""
    return f"${amount:,.2f}" if amount is not None else "—"


def _format_confidence(conf: float | None) -> str:
    """Format 0–1 float as percentage string, or '—' for None."""
    try:
        return f"{float(conf) * 100:.0f}%" if conf is not None else "—"
    except (TypeError, ValueError):
        return "—"


def _sum_amounts(records: list[dict[str, Any]]) -> float:
    """Sum all non-null amount values across records."""
    return sum(float(r["amount"]) for r in records if r.get("amount") is not None)


def _to_display_df(records: list[dict[str, Any]]) -> pd.DataFrame:
    """
    Convert extraction records to a display-ready DataFrame.
    Selects only public columns (drops internal _source / _source_page keys).
    Formats amount and confidence for readability.
    """
    rows = [{k: rec.get(k) for k in _COLUMN_MAP} for rec in records]
    df = pd.DataFrame(rows, columns=list(_COLUMN_MAP.keys()))
    df.rename(columns=_COLUMN_MAP, inplace=True)

    df["Amount ($)"]  = df["Amount ($)"].apply(
        lambda x: _format_currency(x) if pd.notna(x) and x is not None else "—"
    )
    df["Confidence"]  = df["Confidence"].apply(
        lambda x: _format_confidence(x) if pd.notna(x) and x is not None else "—"
    )
    df.fillna("—", inplace=True)
    return df


# ─── Excel Generator ──────────────────────────────────────────────────────────

def generate_excel(records: list[dict[str, Any]]) -> bytes:
    """
    Generate an xlsx workbook with two sheets:
      "Extracted Data" — full transaction table with formatting
      "Summary"        — aggregated metrics for quick overview

    Args:
        records: List of normalised extraction record dicts from extractor.py.

    Returns:
        Raw bytes of the xlsx file, ready for st.download_button.
    """
    df = _to_display_df(records)
    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:

        # ── Sheet 1: Transaction data ──────────────────────────────────────
        df.to_excel(writer, sheet_name="Extracted Data", index=False)
        ws = writer.sheets["Extracted Data"]

        # Style header row (row 1 in Excel = index 1 in openpyxl)
        header_fill = PatternFill("solid", fgColor=_XL_HEADER_HEX)
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Auto-size each column based on max content length (capped at 45)
        for col_cells in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value else 0) for cell in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 45)

        # Apply per-row GST colouring and confidence highlighting (data rows start at 2)
        conf_col_idx = list(_COLUMN_MAP.keys()).index("confidence") + 1  # 1-based

        for excel_row, rec in enumerate(records, start=2):
            gst_val = rec.get("gst_code", "unknown")
            fill    = _XL_GST_FILLS.get(gst_val)
            if fill:
                for cell in ws[excel_row]:
                    cell.fill = fill

            # Red bold for confidence < 85%
            conf = float(rec.get("confidence") or 0)
            if conf < 0.85:
                conf_cell = ws.cell(row=excel_row, column=conf_col_idx)
                conf_cell.font = Font(color="C0392B", bold=True, size=10)

        # ── Sheet 2: Summary metrics ───────────────────────────────────────
        total = _sum_amounts(records)
        summary_rows = [
            ("Total Records",                  len(records)),
            ("Total Amount (incl. GST)",        f"${total:,.2f}"),
            ("Records — 10% GST (Taxable)",    sum(1 for r in records if r.get("gst_code") == "10%")),
            ("Records — 0% GST (Free)",        sum(1 for r in records if r.get("gst_code") == "0%")),
            ("Records — Unknown GST",           sum(1 for r in records if r.get("gst_code") == "unknown")),
            ("Low Confidence (<85%)",           sum(1 for r in records if float(r.get("confidence") or 0) < 0.85)),
            ("Generated At",                    datetime.now().strftime("%d/%m/%Y %H:%M")),
        ]
        summary_df = pd.DataFrame(summary_rows, columns=["Metric", "Value"])
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        ws_s = writer.sheets["Summary"]
        for cell in ws_s[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor=_XL_HEADER_HEX)
        ws_s.column_dimensions["A"].width = 38
        ws_s.column_dimensions["B"].width = 22

    return buf.getvalue()


# ─── PDF Report Generator ─────────────────────────────────────────────────────

def generate_pdf_report(records: list[dict[str, Any]]) -> bytes:
    """
    Generate a professional landscape A4 PDF report using ReportLab Platypus.

    Structure:
      - Title + generated-at metadata line
      - GST summary tile table (4 colour-coded cells)
      - Full transaction detail table with Paragraph cells for word wrapping,
        GST-based row colouring, and red-bold for low-confidence values
      - Disclaimer footer

    Args:
        records: List of normalised extraction record dicts from extractor.py.

    Returns:
        Raw bytes of the PDF file, ready for st.download_button.
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )

    styles = getSampleStyleSheet()

    # ── Custom styles ─────────────────────────────────────────────────────
    title_style = ParagraphStyle(
        "SmTitle",
        parent=styles["Heading1"],
        fontSize=17,
        textColor=_HEADER_BLUE,
        fontName="Helvetica-Bold",
        spaceAfter=3,
    )
    subtitle_style = ParagraphStyle(
        "SmSubtitle",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.grey,
        spaceAfter=10,
    )
    small_cell = ParagraphStyle(
        "SmCell",
        parent=styles["Normal"],
        fontSize=7.5,
        leading=10,       # Line height — keeps multi-line cells compact
        wordWrap="CJK",   # Aggressive word-wrap (breaks within long tokens if needed)
    )
    small_cell_red = ParagraphStyle(
        "SmCellRed",
        parent=small_cell,
        textColor=colors.Color(0.75, 0.05, 0.05),
        fontName="Helvetica-Bold",
    )
    footer_style = ParagraphStyle(
        "SmFooter",
        parent=styles["Normal"],
        fontSize=6.5,
        textColor=colors.grey,
    )

    story: list = []
    total = _sum_amounts(records)

    # ── Title ─────────────────────────────────────────────────────────────
    story.append(Paragraph("Smartify Plus — Financial Extraction Report", title_style))
    story.append(Paragraph(
        f"Generated: {datetime.now().strftime('%d/%m/%Y at %H:%M')}  ·  "
        f"Records: {len(records)}  ·  "
        f"Total Amount: {_format_currency(total)}",
        subtitle_style,
    ))
    story.append(Spacer(1, 4 * mm))

    # ── GST summary tiles ─────────────────────────────────────────────────
    n_10  = sum(1 for r in records if r.get("gst_code") == "10%")
    n_0   = sum(1 for r in records if r.get("gst_code") == "0%")
    n_unk = sum(1 for r in records if r.get("gst_code") == "unknown")
    n_low = sum(1 for r in records if float(r.get("confidence") or 0) < 0.85)

    hdr_para = lambda t: Paragraph(t, ParagraphStyle(  # noqa: E731
        "SumHdr", parent=styles["Normal"], fontSize=8,
        fontName="Helvetica-Bold", textColor=colors.white, alignment=1,
    ))
    val_para = lambda t: Paragraph(t, ParagraphStyle(  # noqa: E731
        "SumVal", parent=styles["Normal"], fontSize=13,
        fontName="Helvetica-Bold", alignment=1,
    ))

    summary_data = [
        [hdr_para("10% GST Taxable"), hdr_para("0% GST Free"),
         hdr_para("Unknown GST"),     hdr_para("Low Confidence (<85%)")],
        [val_para(str(n_10)),         val_para(str(n_0)),
         val_para(str(n_unk)),        val_para(str(n_low))],
    ]
    summary_table = Table(summary_data, colWidths=[55 * mm] * 4)
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), _HEADER_BLUE),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        # Colour-code value row cells by category
        ("BACKGROUND", (0, 1), (0, 1), _GST_COLOURS["10%"]),
        ("BACKGROUND", (1, 1), (1, 1), _GST_COLOURS["0%"]),
        ("BACKGROUND", (2, 1), (2, 1), _GST_COLOURS["unknown"]),
        ("BACKGROUND", (3, 1), (3, 1), colors.Color(1.0, 0.93, 0.75)),  # Amber for low conf
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 6 * mm))

    # ── Transaction detail table ──────────────────────────────────────────
    story.append(Paragraph("Transaction Detail", styles["Heading2"]))
    story.append(Spacer(1, 2 * mm))

    hdr_cell = lambda t: Paragraph(t, ParagraphStyle(  # noqa: E731
        "TblHdr", parent=styles["Normal"], fontSize=8,
        fontName="Helvetica-Bold", textColor=colors.white,
    ))

    headers = ["Date", "Vendor", "Description", "Amount", "GST", "Conf.", "Notes"]
    col_widths = [22*mm, 42*mm, 62*mm, 24*mm, 16*mm, 16*mm, 55*mm]

    table_data: list[list] = [[hdr_cell(h) for h in headers]]

    row_styles: list[tuple] = [
        ("BACKGROUND",   (0, 0), (-1, 0), _HEADER_BLUE),
        ("FONTSIZE",     (0, 0), (-1, -1), 7.5),
        ("GRID",         (0, 0), (-1, -1), 0.3, colors.grey),
        ("VALIGN",       (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING",   (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
    ]

    for row_idx, rec in enumerate(records, start=1):
        conf     = float(rec.get("confidence") or 0)
        gst_code = rec.get("gst_code", "unknown")

        # Use Paragraph objects so long text wraps within cell bounds
        row = [
            Paragraph(str(rec.get("date")        or "—"), small_cell),
            Paragraph(str(rec.get("vendor")       or "—"), small_cell),
            Paragraph(str(rec.get("description")  or "—"), small_cell),
            Paragraph(_format_currency(rec.get("amount")),  small_cell),
            Paragraph(gst_code,                             small_cell),
            # Red bold confidence when below 85% threshold
            Paragraph(_format_confidence(conf), small_cell_red if conf < 0.85 else small_cell),
            Paragraph(str(rec.get("notes") or ""),          small_cell),
        ]
        table_data.append(row)

        # GST-based row background colour
        bg = _GST_COLOURS.get(gst_code, colors.white)
        row_styles.append(("BACKGROUND", (0, row_idx), (-1, row_idx), bg))

    tx_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    tx_table.setStyle(TableStyle(row_styles))
    story.append(tx_table)

    # ── Disclaimer footer ─────────────────────────────────────────────────
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph(
        "DISCLAIMER: This report is AI-generated and must be reviewed by a qualified accountant. "
        "GST classifications are indicative only and do not constitute tax advice. "
        "Always verify all figures against original source documents before lodging with the ATO. "
        "Smartify Plus Phase Zero — For demonstration purposes only.",
        footer_style,
    ))

    doc.build(story)
    return buf.getvalue()
